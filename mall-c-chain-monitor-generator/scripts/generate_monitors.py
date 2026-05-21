#!/usr/bin/env python3
import argparse
import copy
import csv
import hashlib
import json
import re
import uuid
from pathlib import Path


DEFAULT_TEMPLATE = Path(__file__).resolve().parents[1] / "references" / "reviewed-4-checker-template.json"

HEADER_ALIASES = {
    "system_name": ["业务系统名称", "业务系统", "系统名称", "system_name", "system"],
    "name": ["业务链路", "关键业务链路", "链路名称", "name", "chain", "chain_name"],
    "business": ["业务域", "业务名称", "告警业务名", "business", "business_name"],
    "routes": ["关键接口", "接口列表", "接口", "routes", "resource", "resources"],
    "p90_ms": ["P90阈值(ms)", "P90阈值", "p90_ms", "p90", "P90"],
    "p99_ms": ["P99阈值(ms)", "P99阈值", "p99_ms", "p99", "P99"],
    "http_threshold": ["HTTP异常率阈值", "HTTP异常状态占比阈值", "http_threshold"],
    "error_threshold": ["APM错误率阈值", "Span错误率阈值", "错误率阈值", "error_threshold"],
}


def make_uuid():
    return str(uuid.uuid4())


def make_sign_id(seed: str) -> str:
    return hashlib.md5(seed.encode("utf-8")).hexdigest()


def canonical_header(value):
    text = str(value or "").strip().replace(" ", "")
    for canonical, aliases in HEADER_ALIASES.items():
        if text in [a.replace(" ", "") for a in aliases]:
            return canonical
    return None


def parse_number(value, field_name):
    if value is None or str(value).strip() == "":
        raise ValueError(f"Missing required numeric field: {field_name}")
    text = str(value).strip().replace("ms", "").replace("毫秒", "").replace(",", "")
    return int(float(text))


def parse_threshold(value, default_value):
    if value is None or str(value).strip() == "":
        return str(default_value)
    text = str(value).strip().replace(">=", "").replace("%", "").strip()
    return text


def split_route_text(value):
    if isinstance(value, list):
        raw_routes = value
    else:
        raw_routes = re.split(r"[\n;；,，]+", str(value or ""))
    routes = []
    for item in raw_routes:
        route = str(item).strip()
        if not route:
            continue
        routes.append(route)
    return routes


def load_json_input(path):
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(data, list):
        return None, data, {}
    if isinstance(data, dict):
        chains = data.get("chains")
        if chains is None:
            raise ValueError("JSON object input must contain a 'chains' array.")
        return data.get("system_name"), chains, data.get("overall", {}) or {}
    raise ValueError("JSON input must be a list or object.")


def load_csv_input(path):
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        header_map = {h: canonical_header(h) for h in reader.fieldnames or []}
        for raw in reader:
            row = {}
            for src, dst in header_map.items():
                if dst:
                    row[dst] = raw.get(src)
            rows.append(row)
    return rows


def load_xlsx_input(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read XLSX input. Install openpyxl or provide CSV/JSON.") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return []
    canonical = [canonical_header(h) for h in headers]
    rows = []
    for values in rows_iter:
        if not values or all(v is None or str(v).strip() == "" for v in values):
            continue
        row = {}
        for key, value in zip(canonical, values):
            if key:
                row[key] = value
        rows.append(row)
    return rows


def normalize_chains(rows, default_system_name=None, default_http_threshold="1", default_error_threshold="1"):
    chains = []
    system_names = []
    for idx, row in enumerate(rows, 1):
        system_name = str(row.get("system_name") or default_system_name or "").strip()
        if system_name:
            system_names.append(system_name)
        name = str(row.get("name") or "").strip()
        business = str(row.get("business") or "").strip()
        routes = split_route_text(row.get("routes"))
        if not name:
            raise ValueError(f"Row {idx} missing business chain name.")
        if not business:
            raise ValueError(f"Row {idx} missing business suffix/business domain.")
        if not routes:
            raise ValueError(f"Row {idx} missing key interfaces.")
        chains.append(
            {
                "name": name,
                "business": business,
                "routes": routes,
                "p90_ms": parse_number(row.get("p90_ms"), f"row {idx} p90_ms"),
                "p99_ms": parse_number(row.get("p99_ms"), f"row {idx} p99_ms"),
                "http_threshold": parse_threshold(row.get("http_threshold"), default_http_threshold),
                "error_threshold": parse_threshold(row.get("error_threshold"), default_error_threshold),
            }
        )
    if default_system_name:
        system_name = default_system_name
    else:
        unique_systems = list(dict.fromkeys(system_names))
        if len(unique_systems) != 1:
            raise ValueError("Input must provide exactly one business system name, or pass --system-name.")
        system_name = unique_systems[0]
    return system_name, chains


def load_input(path, system_name=None, default_http_threshold="1", default_error_threshold="1"):
    suffix = path.suffix.lower()
    overall = {}
    if suffix == ".json":
        json_system, rows, overall = load_json_input(path)
        system_name, chains = normalize_chains(rows, system_name or json_system, default_http_threshold, default_error_threshold)
    elif suffix == ".csv":
        rows = load_csv_input(path)
        system_name, chains = normalize_chains(rows, system_name, default_http_threshold, default_error_threshold)
    elif suffix in (".xlsx", ".xlsm"):
        rows = load_xlsx_input(path)
        system_name, chains = normalize_chains(rows, system_name, default_http_threshold, default_error_threshold)
    else:
        raise ValueError("Input must be .xlsx, .csv, or .json")
    return system_name, chains, overall


def with_api_prefix(route: str, add_api_prefix=True) -> str:
    if not add_api_prefix:
        return route
    return route if route.startswith("/api/") else f"/api{route}"


def dynamic_prefix(route):
    if "{" in route and "}" in route:
        return route.split("{", 1)[0]
    if "<" in route and ">" in route:
        return route.split("<", 1)[0]
    match = re.search(r"/:[^/]+", route)
    if match:
        return route[: match.start() + 1]
    return None


def split_routes(routes, add_api_prefix=True):
    static_routes = []
    dynamic_prefixes = []
    mappings = []
    for route in routes:
        full_route = with_api_prefix(route, add_api_prefix)
        prefix = dynamic_prefix(full_route)
        if prefix:
            if prefix not in dynamic_prefixes:
                dynamic_prefixes.append(prefix)
            mappings.append((route, full_route, f"match:{prefix}"))
        else:
            if full_route not in static_routes:
                static_routes.append(full_route)
            mappings.append((route, full_route, "in"))
    return static_routes, dynamic_prefixes, mappings


def routes_literal(static_routes):
    return "[" + ", ".join(f"'{route}'" for route in static_routes) + "]"


def resource_condition(static_routes, dynamic_prefixes):
    parts = []
    if static_routes:
        parts.append(f"`resource` IN {routes_literal(static_routes)}")
    parts.extend(f"`resource` = match('{prefix}')" for prefix in dynamic_prefixes)
    if not parts:
        raise ValueError("Each scope must provide at least one route.")
    if len(parts) == 1:
        return parts[0]
    return f"( {' or '.join(parts)} )"


def resource_filters(filter_type: str, static_routes, dynamic_prefixes):
    filters = []
    if static_routes:
        filters.append(
            {
                "id": make_uuid(),
                "op": "in",
                "name": "resource",
                "type": filter_type,
                "logic": "and",
                "value": "",
                "values": static_routes,
            }
        )
    for prefix in dynamic_prefixes:
        filters.append(
            {
                "id": make_uuid(),
                "op": "match",
                "name": "resource",
                "type": filter_type,
                "logic": "or",
                "value": prefix,
                "values": [],
            }
        )
    return filters


def http_error_a_query(condition):
    return (
        "T::RE(`.*`):(count_distinct(`trace_id`)) "
        f"{{ `http_status_group` IN ['4xx', '5xx'] and {condition} }} "
        "BY `project`, `env`, `service`, `resource`, `http_status_code`"
    )


def http_error_b_query(condition):
    return (
        "T::RE(`.*`):(count_distinct(`trace_id`)) "
        f"{{ {condition} }} "
        "BY `project`, `env`, `service`, `resource`, `http_status_code`"
    )


def http_error_expr_query(condition):
    return f'eval(a/b, a="{http_error_a_query(condition)}", b="{http_error_b_query(condition)}")'


def error_rate_query(condition):
    return (
        'eval(A.a1/B.b1,'
        'A="T::RE(`.*`):(COUNT_DISTINCT(`trace_id`) as a1) '
        f"{{`status` = 'error', {condition} }}  "
        'BY `project`, `env`, `service`, `resource`",'
        'B="T::RE(`.*`):(COUNT_DISTINCT(`trace_id`) as b1) '
        f"{{ {condition} }}  "
        'BY `project`, `env`, `service`, `resource`"'
        ").fill(0)"
    )


def error_rate_window_query(condition):
    return (
        'eval(A.a1/B.b1,'
        'A="<window>T::RE(`.*`):(COUNT_DISTINCT(`trace_id`) as a1) '
        f"{{`status` = 'error', {condition} }}  "
        'BY `project`, `env`, `service`, `resource`</window>",'
        'B="<window>T::RE(`.*`):(COUNT_DISTINCT(`trace_id`) as b1) '
        f"{{ {condition} }}  "
        'BY `project`, `env`, `service`, `resource`</window>"'
        ").fill(0)"
    )


def latency_query(condition, percentile):
    return (
        f'eval(A/1000, A="T::RE(`.*`):(percentile(`duration`, {percentile})) '
        f"{{ {condition} }} "
        'BY `project`, `env`, `service`, `resource`")'
    )


def latency_child_query(condition, percentile):
    return (
        f"T::RE(`.*`):(percentile(`duration`, {percentile})) "
        f"{{ {condition} }} "
        "BY `project`, `env`, `service`, `resource`"
    )


def classify_template_checkers(checkers):
    buckets = {}
    for checker in checkers:
        title = checker.get("jsonScript", {}).get("title", "")
        checker_type = checker.get("jsonScript", {}).get("type", "")
        if "http_status_code" in title or "状态码" in title:
            buckets["http"] = checker
        elif "P99" in title:
            buckets["p99"] = checker
        elif "P90" in title:
            buckets["p90"] = checker
        elif checker_type == "apmCheck":
            buckets["error"] = checker
    required = {"http", "error", "p99", "p90"}
    missing = required - set(buckets)
    if missing:
        raise ValueError(f"Template is missing checker types: {sorted(missing)}")
    return buckets


def set_all_rule_thresholds(checker, threshold, match_times=None):
    for rule in checker.get("jsonScript", {}).get("checkerOpt", {}).get("rules", []):
        for condition in rule.get("conditions", []):
            condition["alias"] = "Result"
            condition["operands"] = [str(threshold)]
            condition["operator"] = ">="
        if match_times is not None:
            rule["matchTimes"] = match_times
    for rule in checker.get("extend", {}).get("rules", []):
        for condition in rule.get("conditions", []):
            condition["alias"] = "Result"
            condition["operands"] = [str(threshold)]
            condition["operator"] = ">="
        if match_times is not None:
            rule["matchTimes"] = match_times


def reset_common(checker, system_name, title):
    checker["jsonScript"]["title"] = title
    checker["tagInfo"] = [{"name": system_name}]
    checker["signId"] = make_sign_id(title)
    checker["monitorName"] = checker.get("monitorName") or "default"
    checker["type"] = "trigger"
    checker["is_disable"] = False


def build_scope_checkers(template_by_kind, scope, system_name, add_api_prefix=True):
    static_routes, dynamic_prefixes, mappings = split_routes(scope["routes"], add_api_prefix)
    condition = resource_condition(static_routes, dynamic_prefixes)
    filters_field = resource_filters("field", static_routes, dynamic_prefixes)
    filters_keyword = resource_filters("keyword", static_routes, dynamic_prefixes)
    filters_empty = resource_filters("", static_routes, dynamic_prefixes)

    http_query_a = http_error_a_query(condition)
    http_query_b = http_error_b_query(condition)
    http_query_expr = http_error_expr_query(condition)
    err_query = error_rate_query(condition)
    err_window_query = error_rate_window_query(condition)
    p99_query = latency_query(condition, 99)
    p99_child = latency_child_query(condition, 99)
    p90_query = latency_query(condition, 90)
    p90_child = latency_child_query(condition, 90)

    suffix = scope["suffix"]
    http_threshold = scope.get("http_threshold", "1")
    error_threshold = scope.get("error_threshold", "1")

    http_checker = copy.deepcopy(template_by_kind["http"])
    reset_common(http_checker, system_name, f"{{{{project}}}}产品服务{{{{service}}}} 接口{{{{resource}}}}响应状态码{{{{http_status_code}}}}异常率告警-{suffix}")
    http_checker["jsonScript"]["targets"][0]["dql"] = http_query_expr
    set_all_rule_thresholds(http_checker, http_threshold)
    ql = http_checker.get("extend", {}).get("querylist", [])
    if len(ql) >= 3:
        ql[0]["uuid"] = make_uuid()
        ql[0]["query"]["q"] = http_query_a
        if ql[0]["query"].get("filters"):
            ql[0]["query"]["filters"][0]["id"] = make_uuid()
            ql[0]["query"]["filters"][0]["type"] = ""
            ql[0]["query"]["filters"][1:] = filters_empty
        ql[1]["uuid"] = make_uuid()
        ql[1]["query"]["q"] = http_query_b
        ql[1]["query"]["filters"] = filters_keyword
        ql[2]["uuid"] = make_uuid()
        ql[2]["query"]["q"] = http_query_expr

    error_checker = copy.deepcopy(template_by_kind["error"])
    reset_common(error_checker, system_name, f"{{{{project}}}}产品服务{{{{service}}}}关键业务接口{{{{resource}}}}异常错误率告警-{suffix}")
    error_checker["jsonScript"]["targets"][0]["dql"] = err_query
    error_checker["jsonScript"]["windowDql"] = err_window_query
    set_all_rule_thresholds(error_checker, error_threshold)
    ql = error_checker.get("extend", {}).get("querylist", [])
    if ql:
        ql[0]["uuid"] = make_uuid()
        ql[0]["query"]["q"] = err_query
        ql[0]["query"]["filters"] = filters_field

    p99_checker = copy.deepcopy(template_by_kind["p99"])
    reset_common(p99_checker, system_name, f"{{{{project}}}}产品服务{{{{service}}}}关键业务接口{{{{resource}}}} P99响应时间连续超过阈值-{suffix}")
    p99_checker["jsonScript"]["targets"][0]["dql"] = p99_query
    set_all_rule_thresholds(p99_checker, scope["p99_ms"])
    ql = p99_checker.get("extend", {}).get("querylist", [])
    if ql:
        ql[0]["uuid"] = make_uuid()
        ql[0]["query"]["q"] = p99_query
        ql[0]["query"]["filters"] = filters_keyword
        if ql[0]["query"].get("children"):
            ql[0]["query"]["children"][0]["q"] = p99_child
            ql[0]["query"]["children"][0]["filters"] = filters_keyword

    p90_checker = copy.deepcopy(template_by_kind["p90"])
    reset_common(p90_checker, system_name, f"{{{{project}}}}产品服务{{{{service}}}}关键业务接口{{{{resource}}}} P90响应时间连续超过阈值-{suffix}")
    p90_checker["jsonScript"]["targets"][0]["dql"] = p90_query
    set_all_rule_thresholds(p90_checker, scope["p90_ms"])
    ql = p90_checker.get("extend", {}).get("querylist", [])
    if ql:
        ql[0]["uuid"] = make_uuid()
        ql[0]["query"]["q"] = p90_query
        ql[0]["query"]["filters"] = filters_keyword
        if ql[0]["query"].get("children"):
            ql[0]["query"]["children"][0]["q"] = p90_child
            ql[0]["query"]["children"][0]["filters"] = filters_keyword

    return [http_checker, error_checker, p99_checker, p90_checker], mappings


def build_overall_scope(system_name, chains, overall_options):
    all_routes = []
    for chain in chains:
        all_routes.extend(chain["routes"])
    return {
        "name": system_name,
        "suffix": system_name,
        "routes": all_routes,
        "p90_ms": parse_number(overall_options.get("p90_ms"), "overall p90_ms") if overall_options.get("p90_ms") else max(c["p90_ms"] for c in chains),
        "p99_ms": parse_number(overall_options.get("p99_ms"), "overall p99_ms") if overall_options.get("p99_ms") else max(c["p99_ms"] for c in chains),
        "http_threshold": parse_threshold(overall_options.get("http_threshold"), chains[0].get("http_threshold", "1")),
        "error_threshold": parse_threshold(overall_options.get("error_threshold"), chains[0].get("error_threshold", "1")),
    }


def write_json(path: Path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_filters_report(path, system_name, chains, scope_mappings):
    if not path:
        return
    lines = [f"# {system_name} 监控器接口筛选清单", ""]
    lines.append(f"- 业务系统：{system_name}")
    lines.append(f"- 业务链路数：{len(chains)}")
    lines.append(f"- 原始接口数：{sum(len(c['routes']) for c in chains)}")
    lines.append("")
    lines.append("| 范围 | 原始接口 | 归一化接口 | 筛选方式 |")
    lines.append("|---|---|---|---|")
    for scope_name, mappings in scope_mappings:
        for raw, normalized, mode in mappings:
            if mode == "in":
                desc = "resource IN 精确匹配"
            else:
                desc = "resource = match 前缀：" + mode.split(":", 1)[1]
            lines.append(f"| {scope_name} | `{raw}` | `{normalized}` | {desc} |")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def write_summary(path, system_name, output_path, chain_count, monitor_count, filters_report):
    if not path:
        return
    text = f"""# {system_name} 关键业务 SLO/SLI 监控器生成摘要

- 业务系统：{system_name}
- 业务链路数：{chain_count}
- 监控器数量：{monitor_count}
- 输出 JSON：`{output_path}`
- 接口筛选清单：`{filters_report}`

## 监控器组成

每条业务链路 4 个监控器，总体业务系统 4 个监控器：

1. HTTP 状态码异常率
2. APM 错误率
3. P99 响应时间
4. P90 响应时间

## 命名规则

总体业务系统监控器使用业务系统名称作为标题后缀；分链路监控器使用业务域作为标题后缀。

## 标签规则

所有监控器仅包含一个标签：`{system_name}`。
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="Generate one Guance monitor JSON for per-chain and overall business-system SLI monitors.")
    parser.add_argument("--input", required=True, help="Customer key interface table: .xlsx, .csv, or .json")
    parser.add_argument("--output", required=True, help="Path to the single output monitor JSON.")
    parser.add_argument("--system-name", help="Business system name. Overrides table/JSON system name.")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="Reviewed 4-checker template JSON. Defaults to bundled template.")
    parser.add_argument("--overall-p90-ms", type=int, help="Overall system P90 threshold. Defaults to max chain P90.")
    parser.add_argument("--overall-p99-ms", type=int, help="Overall system P99 threshold. Defaults to max chain P99.")
    parser.add_argument("--http-threshold", default="1", help="Default HTTP anomaly threshold. Defaults to 1 to match reviewed template style.")
    parser.add_argument("--error-threshold", default="1", help="Default APM error threshold. Defaults to 1 to match reviewed template style.")
    parser.add_argument("--no-api-prefix", action="store_true", help="Do not prepend /api to routes.")
    parser.add_argument("--filters-report", help="Optional Markdown path for route filter mapping report.")
    parser.add_argument("--summary", help="Optional Markdown generation summary path.")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    template_path = Path(args.template)
    template = json.loads(template_path.read_text(encoding="utf-8-sig"))
    if not isinstance(template, dict) or "checkers" not in template:
        raise ValueError("Template JSON must be an object with a 'checkers' array.")

    system_name, chains, overall_options = load_input(input_path, args.system_name, args.http_threshold, args.error_threshold)
    if args.overall_p90_ms is not None:
        overall_options["p90_ms"] = args.overall_p90_ms
    if args.overall_p99_ms is not None:
        overall_options["p99_ms"] = args.overall_p99_ms
    if args.http_threshold:
        overall_options.setdefault("http_threshold", args.http_threshold)
    if args.error_threshold:
        overall_options.setdefault("error_threshold", args.error_threshold)

    template_by_kind = classify_template_checkers(template["checkers"])
    combined = {"checkers": []}
    scope_mappings = []
    add_api_prefix = not args.no_api_prefix

    for chain in chains:
        scope = {
            "name": chain["name"],
            "suffix": chain["business"],
            "routes": chain["routes"],
            "p90_ms": chain["p90_ms"],
            "p99_ms": chain["p99_ms"],
            "http_threshold": chain.get("http_threshold", args.http_threshold),
            "error_threshold": chain.get("error_threshold", args.error_threshold),
        }
        checkers, mappings = build_scope_checkers(template_by_kind, scope, system_name, add_api_prefix)
        combined["checkers"].extend(checkers)
        scope_mappings.append((chain["name"], mappings))

    overall_scope = build_overall_scope(system_name, chains, overall_options)
    overall_checkers, overall_mappings = build_scope_checkers(template_by_kind, overall_scope, system_name, add_api_prefix)
    combined["checkers"].extend(overall_checkers)
    scope_mappings.append((f"总体业务系统：{system_name}", overall_mappings))

    write_json(output_path, combined)
    filters_report = Path(args.filters_report) if args.filters_report else None
    summary_path = Path(args.summary) if args.summary else None
    write_filters_report(filters_report, system_name, chains, scope_mappings)
    write_summary(summary_path, system_name, output_path, len(chains), len(combined["checkers"]), filters_report)
    print(f"output={output_path}")
    print(f"system_name={system_name}")
    print(f"chains={len(chains)}")
    print(f"checkers={len(combined['checkers'])}")


if __name__ == "__main__":
    main()

